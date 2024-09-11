"""
    main.py
    Cayden Garcia
    Fall 2024 - Advanced Software Engineering
    Will call all main functions of SlackR
"""
"""Libraries"""
from selenium import webdriver #Selenium -> webdriver: Used to connect web driver to program
from selenium.webdriver.chrome.service import Service #Selenium -> Service
from selenium.webdriver.chrome.options import Options #Selenium -> Options: Allows for custom options to be called
from selenium.webdriver.common.by import By #Selenium -> By: Allows for calls by specific variables from html code
from webdriver_manager.chrome import ChromeDriverManager #Webdriver_manager -> ChromeDriverManager: Allows for chrome driver to be ran
import openpyxl #openpyxl: Allows for storage of data and creation of .xlsx sheet 

"""Variables"""
header_list = ["Name","KD", "Win %","Top Agent","HS %"] # List to store all headers I want, 
                                            #just make sure the data being pulled is in the same position as the list position

data_list = [] # Initialize empty list to store data
row_ = 1 #integer for accesing specific row
column_ = 1 #integer for accessing specific 


"""Selenium"""
options = Options() #intialize option variable to class Options
options.add_argument('--headless') #Take away window from being open on program run

driver = webdriver.Chrome(service = Service(ChromeDriverManager().install()), options = options) #Initializes driver var to chrome driver
driver.get("https://valorantstats.xyz/stats/profile/SEN%20curry-lisa?actId=all&gameMode=all") #URL for what html page I want
driver.implicitly_wait(10) #Makes driver wait 10 ms before doing anything: Allows for everything to load before accessing HTML elements



name_finder = driver.find_element(By.CLASS_NAME, 'user-name') #Finds username ***for now***
name = name_finder.text #Stores name

kd_finder = driver.find_elements(By.CLASS_NAME,'ov-stat-value') #Finds where KD is
kd = kd_finder[0] #Limits element to KD
kd_number = kd.text #Stores KD

winp_finder = driver.find_elements(By.CLASS_NAME,'ov-stat-value') #Finds where win% is stored
winp = winp_finder[1] #Limits element to win%
winp_number = winp.text #stores win%

agent_finder = driver.find_elements(By.CLASS_NAME, 'info-name') #Finds Top 3 agents
top_agent = agent_finder[0] #Limits it to top agent
top_agent_name = top_agent.text #stores top agent

hs_percentage_finder = driver.find_elements(By.CLASS_NAME,'ov-stat-value') #Finds where HS% is stored
hs_percentage = hs_percentage_finder[2] #Limits element to HS%
hs_percentage_number = hs_percentage.text #stores HS%

#Adding all calculated stats to data_list
data_list.append(name) #Add name to list
data_list.append(kd_number) #Add kd to list
data_list.append(winp_number) #Add win% to list
data_list.append(top_agent_name) #Add top agent to list
data_list.append(hs_percentage_number) #Add HS% to list



driver.quit() #Once all data is taken and stored into data_list, quit the driver


"""pyxl"""
workbook = openpyxl.Workbook() #opens workbook 
sheet = workbook.active #sets workbook to active
#for loop to store data into a new column on a single row, must enumerate header_list since 'i' refers to each element not the index
for i, header in enumerate(header_list):
    sheet.cell(row = row_, column = column_, value = header_list[i])
    column_ += 1
#storing data into spots(CREATE FOR LOOP)  
sheet.cell(row=2,column=1,value = data_list[0]) #Store name into (2,1)
sheet.cell(row=2,column=2,value = data_list[1]) #Store KD into (2,2)
sheet.cell(row=2,column=3,value = data_list[2]) #Store Win% into (2,3)
sheet.cell(row=2,column=4,value = data_list[3]) #Store Top agent into (2,4)
sheet.cell(row=2,column=5,value = data_list[4]) #Store Top agent into (2,4)

workbook.save("SlackR_stats.xlsx") #Saves file to workbook
print("Data has been saved to SlackR_stats.xlsx") #Ouput to show program is done running